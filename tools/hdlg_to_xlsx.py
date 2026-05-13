#!/usr/bin/env python3
"""Export an HLZS-wrapped HDGL dialog table to an Excel workbook."""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from hdlg_codec import read_hdlg


HEADERS = [
    "id",
    "ja_original",
    "text",
    "en_reference",
    "notes",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export script_dialog_*.hdlg strings to .xlsx."
    )
    parser.add_argument("input_hdlg", help="source .hdlg file")
    parser.add_argument("output_xlsx", help="output .xlsx file")
    parser.add_argument(
        "--reference",
        help="optional reference .hdlg, usually script_dialog_en.hdlg",
    )
    return parser.parse_args()


def style_dialog_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 10,
        "B": 54,
        "C": 54,
        "D": 54,
        "E": 28,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column in (2, 3, 4, 5):
                cell.number_format = "@"


def write_meta_sheet(wb, source: Path, output: Path, hlzs_version: int, table) -> None:
    ws = wb.create_sheet("meta")
    rows = [
        ("source_file", str(source)),
        ("output_file", str(output)),
        ("format", "HLZS/HDGL"),
        ("hlzs_version", f"0x{hlzs_version:08X}"),
        ("hdgl_version", f"0x{table.version:08X}"),
        ("entry_count", table.count),
        ("data_offset", f"0x{table.data_offset:08X}"),
        ("encoding", "utf-8"),
        ("trailing_size", table.trailing_size),
        ("exported_at", datetime.now().isoformat(timespec="seconds")),
    ]

    ws.append(["key", "value"])
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80


def export_xlsx(input_hdlg: Path, output_xlsx: Path, reference: Path | None) -> None:
    _, hlzs_version, table = read_hdlg(input_hdlg)
    reference_entries = None

    if reference:
        _, _, reference_table = read_hdlg(reference)
        if reference_table.count != table.count:
            raise SystemExit(
                "reference entry count mismatch: "
                f"{reference_table.count} != {table.count}"
            )
        reference_entries = reference_table.entries

    wb = Workbook()
    ws = wb.active
    ws.title = "dialog"
    ws.append(HEADERS)

    for entry in table.entries:
        en_reference = ""
        if reference_entries is not None:
            en_reference = reference_entries[entry.index].text
        ws.append(
            [
                entry.index,
                entry.text,
                "",
                en_reference,
                "",
            ]
        )

    style_dialog_sheet(ws)
    write_meta_sheet(wb, input_hdlg, output_xlsx, hlzs_version, table)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def main() -> None:
    args = parse_args()
    export_xlsx(
        Path(args.input_hdlg),
        Path(args.output_xlsx),
        Path(args.reference) if args.reference else None,
    )
    print(f"exported: {args.output_xlsx}")


if __name__ == "__main__":
    main()
