#!/usr/bin/env python3
"""Build an HLZS-wrapped HDGL dialog table from an Excel workbook."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from hdlg_codec import HDGL_VERSION, HLZS_VERSION, HdlgFormatError, write_hdlg


REQUIRED_COLUMNS = ["id", "text"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build .hdlg from an .xlsx workbook.")
    parser.add_argument("input_xlsx", help="input .xlsx file")
    parser.add_argument("output_hdlg", help="output .hdlg file")
    parser.add_argument(
        "--sheet",
        default="dialog",
        help="worksheet name containing dialog rows (default: dialog)",
    )
    parser.add_argument(
        "--text-column",
        default="text",
        help="column header to build from (default: text)",
    )
    parser.add_argument(
        "--allow-empty",
        action="store_true",
        help=(
            "build blank text cells as empty strings instead of falling back to "
            "ja_original"
        ),
    )
    parser.add_argument(
        "--hdgl-version",
        default=f"0x{HDGL_VERSION:08X}",
        help="HDGL version to write (default: 0x00010000)",
    )
    parser.add_argument(
        "--hlzs-version",
        default=f"0x{HLZS_VERSION:08X}",
        help="HLZS version to write (default: 0x00001000)",
    )
    return parser.parse_args()


def parse_int(value: str) -> int:
    return int(str(value), 0)


def load_meta_count(wb) -> int | None:
    if "meta" not in wb.sheetnames:
        return None

    ws = wb["meta"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        key, value = row[0], row[1]
        if key == "entry_count":
            return int(value)
    return None


def header_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.column
    return headers


def is_blank_cell_value(value) -> bool:
    return value is None or value == ""


def read_texts(
    workbook_path: Path,
    sheet_name: str,
    text_column: str,
    allow_empty: bool,
) -> list[str]:
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"sheet not found: {sheet_name}")

    ws = wb[sheet_name]
    headers = header_map(ws)
    missing = [name for name in ["id", text_column] if name not in headers]
    if missing:
        raise SystemExit(f"missing required column(s): {', '.join(missing)}")

    id_col = headers["id"]
    text_col = headers[text_column]
    fallback_col = headers.get("ja_original")
    texts_by_id: dict[int, str] = {}

    for row_index in range(2, ws.max_row + 1):
        id_cell = ws.cell(row=row_index, column=id_col)
        text_cell = ws.cell(row=row_index, column=text_col)

        if id_cell.value is None:
            continue
        try:
            dialog_id = int(id_cell.value)
        except (TypeError, ValueError) as exc:
            raise SystemExit(f"row {row_index}: invalid id {id_cell.value!r}") from exc

        if dialog_id in texts_by_id:
            raise SystemExit(f"duplicate id: {dialog_id}")

        if text_cell.data_type == "f":
            raise SystemExit(
                f"row {row_index}, id {dialog_id}: formulas are not allowed in text cells"
            )

        value = text_cell.value
        if is_blank_cell_value(value):
            if allow_empty:
                value = ""
            elif fallback_col is not None:
                fallback_cell = ws.cell(row=row_index, column=fallback_col)
                if fallback_cell.data_type == "f":
                    raise SystemExit(
                        f"row {row_index}, id {dialog_id}: formulas are not allowed "
                        "in ja_original fallback cells"
                    )
                value = fallback_cell.value
                if is_blank_cell_value(value):
                    raise SystemExit(
                        f"row {row_index}, id {dialog_id}: both text and ja_original "
                        "are blank"
                    )
            else:
                raise SystemExit(
                    f"row {row_index}, id {dialog_id}: blank text cell and no "
                    "ja_original fallback column"
                )

        text = str(value)
        if "\0" in text:
            raise SystemExit(f"row {row_index}, id {dialog_id}: NUL byte is not allowed")

        texts_by_id[dialog_id] = text

    if not texts_by_id:
        raise SystemExit("no dialog rows found")

    expected_count = load_meta_count(wb)
    if expected_count is None:
        expected_count = max(texts_by_id) + 1

    expected_ids = set(range(expected_count))
    actual_ids = set(texts_by_id)
    if actual_ids != expected_ids:
        missing = sorted(expected_ids - actual_ids)[:20]
        extra = sorted(actual_ids - expected_ids)[:20]
        detail = []
        if missing:
            detail.append(f"missing ids: {missing}")
        if extra:
            detail.append(f"unexpected ids: {extra}")
        raise SystemExit("; ".join(detail))

    return [texts_by_id[index] for index in range(expected_count)]


def main() -> None:
    args = parse_args()
    texts = read_texts(
        Path(args.input_xlsx),
        args.sheet,
        args.text_column,
        args.allow_empty,
    )

    output = Path(args.output_hdlg)
    output.parent.mkdir(parents=True, exist_ok=True)
    try:
        write_hdlg(
            texts,
            output,
            hdgl_version=parse_int(args.hdgl_version),
            hlzs_version=parse_int(args.hlzs_version),
        )
    except HdlgFormatError as exc:
        raise SystemExit(str(exc)) from exc

    print(f"built: {args.output_hdlg}")
    print(f"entries: {len(texts)}")


if __name__ == "__main__":
    main()
